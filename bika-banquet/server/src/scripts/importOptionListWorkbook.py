import json
import sys
from openpyxl import load_workbook


def normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: importOptionListWorkbook.py <workbook-path> [sheet-name ...]")

    workbook_path = sys.argv[1]
    target_sheets = sys.argv[2:] or ["Item List", "Item List(2)"]
    workbook = load_workbook(workbook_path, data_only=True)
    payload = {"sheets": []}

    for sheet_name in target_sheets:
      sheet = workbook[sheet_name]
      rows = []
      for row in sheet.iter_rows(values_only=True):
          rows.append([normalize_cell(cell) for cell in row])
      payload["sheets"].append({"name": sheet_name, "rows": rows})

    print(json.dumps(payload))


if __name__ == "__main__":
    main()
